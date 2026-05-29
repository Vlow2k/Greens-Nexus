import csv
import io
import re
from datetime import datetime
from fastapi import FastAPI, Depends, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional
import models
from database import engine, get_db
from unifi_client import fetch_all, build_site_payload

models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="Greens Nexus API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        "https://vlow2k.github.io",
        "https://nexus.greensglobal.com",
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
def health():
    return {"status": "ok"}


# ── Pydantic Schemas ────────────────────────────────────────────────────────

class TaskCreate(BaseModel):
    id: str
    title: str
    assignee: str
    project: str
    due_date: str
    hours: str
    comment: str = ""
    priority: str
    status: str
    dept: str
    synced: bool = True

class TaskUpdate(BaseModel):
    status: Optional[str] = None
    comment: Optional[str] = None

class PurchaseCreate(BaseModel):
    item: str
    vendor: str = ""
    cost: float = 0
    qty: int = 1
    dept: str
    status: str = "pending"

class PurchaseStatusUpdate(BaseModel):
    status: str

class ReviewReply(BaseModel):
    reply_text: str

class AssetCreate(BaseModel):
    name: str
    category: str
    assigned_to: str = "Unassigned"
    status: str = "Available"
    last_seen: str

class UserCreate(BaseModel):
    name: str
    dept: str
    role: str
    access_level: str
    status: str = "Active"
    last_login: str = ""

class WebsiteCreate(BaseModel):
    name: str
    domain: str
    ssl_days: int = 90
    uptime: float = 99.9
    status: str = "Online"

class ExternalLinkCreate(BaseModel):
    name: str
    url: str
    category: str
    description: str = ""

class SopCreate(BaseModel):
    title: str
    category: str
    status: str = "Published"
    date: str

class OpsProjectCreate(BaseModel):
    name: str
    status: str = "on-track"
    location: str
    members: int = 0
    due_date: str
    progress: int = 0


# ── Tasks ───────────────────────────────────────────────────────────────────

@app.get("/tasks")
def list_tasks(db: Session = Depends(get_db)):
    return db.query(models.Task).all()

@app.post("/tasks", status_code=201)
def create_task(task: TaskCreate, db: Session = Depends(get_db)):
    db_task = models.Task(**task.model_dump())
    db.add(db_task)
    db.commit()
    db.refresh(db_task)
    return db_task

@app.patch("/tasks/{task_id}")
def update_task(task_id: str, update: TaskUpdate, db: Session = Depends(get_db)):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    if update.status is not None:
        task.status = update.status
    if update.comment is not None:
        task.comment = update.comment
    db.commit()
    db.refresh(task)
    return task

@app.delete("/tasks/{task_id}", status_code=204)
def delete_task(task_id: str, db: Session = Depends(get_db)):
    task = db.query(models.Task).filter(models.Task.id == task_id).first()
    if not task:
        raise HTTPException(status_code=404, detail="Task not found")
    db.delete(task)
    db.commit()


# ── Purchase Requests ───────────────────────────────────────────────────────

@app.get("/purchase-requests")
def list_purchase_requests(db: Session = Depends(get_db)):
    return db.query(models.PurchaseRequest).all()

@app.post("/purchase-requests", status_code=201)
def create_purchase_request(req: PurchaseCreate, db: Session = Depends(get_db)):
    db_req = models.PurchaseRequest(**req.model_dump())
    db.add(db_req)
    db.commit()
    db.refresh(db_req)
    return db_req

@app.patch("/purchase-requests/{req_id}")
def update_purchase_status(req_id: int, update: PurchaseStatusUpdate, db: Session = Depends(get_db)):
    req = db.query(models.PurchaseRequest).filter(models.PurchaseRequest.id == req_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    req.status = update.status
    db.commit()
    db.refresh(req)
    return req


# ── Reviews ─────────────────────────────────────────────────────────────────

@app.get("/reviews")
def list_reviews(db: Session = Depends(get_db)):
    return db.query(models.Review).all()

@app.patch("/reviews/{review_id}/reply")
def reply_to_review(review_id: int, reply: ReviewReply, db: Session = Depends(get_db)):
    review = db.query(models.Review).filter(models.Review.id == review_id).first()
    if not review:
        raise HTTPException(status_code=404, detail="Review not found")
    review.replied = True
    review.reply_text = reply.reply_text
    db.commit()
    db.refresh(review)
    return review


# ── Marketing Campaigns ─────────────────────────────────────────────────────

@app.get("/marketing-campaigns")
def list_campaigns(db: Session = Depends(get_db)):
    return db.query(models.MarketingCampaign).all()


# ── SOP Updates ─────────────────────────────────────────────────────────────

@app.get("/sop-updates")
def list_sops(db: Session = Depends(get_db)):
    return db.query(models.SopUpdate).all()

@app.post("/sop-updates", status_code=201)
def create_sop(sop: SopCreate, db: Session = Depends(get_db)):
    db_sop = models.SopUpdate(**sop.model_dump())
    db.add(db_sop)
    db.commit()
    db.refresh(db_sop)
    return db_sop


# ── Assets ───────────────────────────────────────────────────────────────────

@app.get("/assets")
def list_assets(db: Session = Depends(get_db)):
    return db.query(models.Asset).all()

@app.post("/assets", status_code=201)
def create_asset(asset: AssetCreate, db: Session = Depends(get_db)):
    db_asset = models.Asset(**asset.model_dump())
    db.add(db_asset)
    db.commit()
    db.refresh(db_asset)
    return db_asset


# ── Users ────────────────────────────────────────────────────────────────────

@app.get("/users")
def list_users(db: Session = Depends(get_db)):
    return db.query(models.User).all()

@app.post("/users", status_code=201)
def create_user(user: UserCreate, db: Session = Depends(get_db)):
    db_user = models.User(**user.model_dump())
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user


# ── Websites ─────────────────────────────────────────────────────────────────

@app.get("/websites")
def list_websites(db: Session = Depends(get_db)):
    return db.query(models.Website).all()

@app.post("/websites", status_code=201)
def create_website(site: WebsiteCreate, db: Session = Depends(get_db)):
    db_site = models.Website(**site.model_dump())
    db.add(db_site)
    db.commit()
    db.refresh(db_site)
    return db_site


# ── External Links ───────────────────────────────────────────────────────────

@app.get("/external-links")
def list_external_links(db: Session = Depends(get_db)):
    return db.query(models.ExternalLink).all()

@app.post("/external-links", status_code=201)
def create_external_link(link: ExternalLinkCreate, db: Session = Depends(get_db)):
    db_link = models.ExternalLink(**link.model_dump())
    db.add(db_link)
    db.commit()
    db.refresh(db_link)
    return db_link

@app.patch("/external-links/{link_id}/click")
def increment_click(link_id: int, db: Session = Depends(get_db)):
    link = db.query(models.ExternalLink).filter(models.ExternalLink.id == link_id).first()
    if not link:
        raise HTTPException(status_code=404, detail="Link not found")
    link.clicks += 1
    db.commit()
    return link


# ── Accounting ───────────────────────────────────────────────────────────────

@app.get("/accounting/transactions")
def list_transactions(db: Session = Depends(get_db)):
    return db.query(models.AccountingTrx).all()

@app.get("/accounting/ramp")
def list_ramp(db: Session = Depends(get_db)):
    return db.query(models.RampTransaction).all()

@app.get("/accounting/ama")
def list_ama(db: Session = Depends(get_db)):
    return db.query(models.AmaEntity).all()


# ── Operations ───────────────────────────────────────────────────────────────

@app.get("/ops-projects")
def list_ops_projects(db: Session = Depends(get_db)):
    return db.query(models.OpsProject).all()

@app.post("/ops-projects", status_code=201)
def create_ops_project(proj: OpsProjectCreate, db: Session = Depends(get_db)):
    db_proj = models.OpsProject(**proj.model_dump())
    db.add(db_proj)
    db.commit()
    db.refresh(db_proj)
    return db_proj


# ── Development ──────────────────────────────────────────────────────────────

@app.get("/dev-projects")
def list_dev_projects(db: Session = Depends(get_db)):
    return db.query(models.DevProject).all()


# ── LMS ──────────────────────────────────────────────────────────────────────

@app.get("/lms-courses")
def list_lms_courses(db: Session = Depends(get_db)):
    return db.query(models.LmsCourse).all()


# ── UniFi ────────────────────────────────────────────────────────────────────

@app.get("/unifi/overview")
async def unifi_overview():
    sites_raw, devices_raw = await fetch_all()
    host_map = {e["hostId"]: e for e in devices_raw.get("data", [])}
    result = [
        build_site_payload(s, host_map.get(s.get("hostId", ""), {}))
        for s in sites_raw.get("data", [])
    ]
    return {"data": result}


@app.get("/unifi/stats")
async def unifi_stats(siteId: str):
    sites_raw, devices_raw = await fetch_all()
    site = next((s for s in sites_raw.get("data", []) if s.get("siteId") == siteId), None)
    if not site:
        raise HTTPException(404, f"Site {siteId} not found")
    host_id = site.get("hostId", "")
    host_entry = next((e for e in devices_raw.get("data", []) if e.get("hostId") == host_id), {})
    payload = build_site_payload(site, host_entry)
    devices = host_entry.get("devices", [])
    return {
        "total_devices": payload["total_devices"],
        "online_devices": payload["online_devices"],
        "offline_devices": len(payload["offline_devices"]),
        "total_clients": payload["wifi_clients"] + payload["wired_clients"],
        "wireless_clients": payload["wifi_clients"],
        "wired_clients": payload["wired_clients"],
        "devices": devices,
    }


@app.get("/unifi/export/csv")
async def unifi_export_csv(siteId: str):
    sites_raw, devices_raw = await fetch_all()
    site = next((s for s in sites_raw.get("data", []) if s.get("siteId") == siteId), None)
    if not site:
        raise HTTPException(404, f"Site {siteId} not found")
    host_id = site.get("hostId", "")
    host_entry = next((e for e in devices_raw.get("data", []) if e.get("hostId") == host_id), {})
    site_name = host_entry.get("hostName") or site.get("meta", {}).get("desc") or siteId
    devices = host_entry.get("devices", [])
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Site", "Name", "Model", "MAC Address", "IP Address", "Status", "Firmware Version", "Firmware Status", "Product Line", "Is Console", "Startup Time", "Adoption Time"])
    for d in devices:
        writer.writerow([site_name, d.get("name", ""), d.get("model", ""), d.get("mac", ""), d.get("ip", ""), d.get("status", ""), d.get("version", ""), d.get("firmwareStatus", ""), d.get("productLine", ""), "Yes" if d.get("isConsole") else "No", d.get("startupTime", ""), d.get("adoptionTime", "")])
    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "_", site_name)
    output.seek(0)
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f"attachment; filename={safe_name}_inventory.csv"})


# ── Dashboard Summary ────────────────────────────────────────────────────────

@app.get("/dashboard/summary")
def dashboard_summary(db: Session = Depends(get_db)):
    tasks_count = db.query(models.Task).filter(models.Task.status != 'Completed').count()
    approvals_count = db.query(models.PurchaseRequest).filter(models.PurchaseRequest.status == 'pending').count()
    req_pending = db.query(models.Requisition).filter(models.Requisition.status == 'pending_manager').count()
    purchases_count = db.query(models.PurchaseRequest).count()
    reviews_pending = db.query(models.Review).filter(models.Review.replied.is_(False)).count()
    sop_count = db.query(models.SopUpdate).count()
    return {
        "tasks_count": tasks_count,
        "approvals_count": approvals_count + req_pending,
        "purchases_count": purchases_count,
        "reviews_pending": reviews_pending,
        "sop_count": sop_count,
    }


# ── Requisitions ─────────────────────────────────────────────────────────────

class RequisitionCreate(BaseModel):
    id: str
    employee_name: str
    employee_dept: str
    item: str
    quantity: int = 1
    reason: str = ""
    status: str = "pending_manager"
    supervisor_name: str = ""

class RequisitionApprove(BaseModel):
    manager_name: str

class RequisitionReject(BaseModel):
    manager_name: str
    rejection_reason: str

class RequisitionAllocate(BaseModel):
    asset_id: str
    supervisor_name: str
    expected_return_date: str = ""

class RequisitionReturn(BaseModel):
    initiated_by: str

class RequisitionConfirmReturn(BaseModel):
    supervisor_name: str
    condition: str = "Available"

class RequisitionMarkLost(BaseModel):
    supervisor_name: str
    notes: str = ""


def _ts():
    return datetime.utcnow().isoformat()


@app.get("/requisitions")
def list_requisitions(db: Session = Depends(get_db)):
    return db.query(models.Requisition).all()

@app.post("/requisitions", status_code=201)
def create_requisition(data: RequisitionCreate, db: Session = Depends(get_db)):
    req = models.Requisition(**data.model_dump(), created_at=_ts(), updated_at=_ts())
    db.add(req); db.commit(); db.refresh(req)
    return req

@app.patch("/requisitions/{req_id}/approve")
def approve_requisition(req_id: str, body: RequisitionApprove, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req: raise HTTPException(404, "Requisition not found")
    req.status = "manager_approved"
    req.manager_name = body.manager_name
    req.manager_approval_date = _ts()
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Approved", action_by=body.manager_name, action_role="Manager", created_at=_ts()))
    db.commit(); db.refresh(req)
    return req

@app.patch("/requisitions/{req_id}/reject")
def reject_requisition(req_id: str, body: RequisitionReject, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req: raise HTTPException(404, "Requisition not found")
    req.status = "rejected"
    req.manager_name = body.manager_name
    req.rejection_reason = body.rejection_reason
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Rejected", action_by=body.manager_name, action_role="Manager", comment=body.rejection_reason, created_at=_ts()))
    db.commit(); db.refresh(req)
    return req

@app.patch("/requisitions/{req_id}/allocate")
def allocate_asset(req_id: str, body: RequisitionAllocate, db: Session = Depends(get_db)):
    req   = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == body.asset_id).first()
    if not req:   raise HTTPException(404, "Requisition not found")
    if not asset: raise HTTPException(404, "Asset not found")
    if asset.status != "Available": raise HTTPException(400, "Asset not available")

    asset.status = "Checked Out"
    asset.assigned_to = req.employee_name
    asset.assigned_req_id = req_id
    asset.last_updated = _ts()[:10]

    req.status = "asset_allocated"
    req.asset_id = body.asset_id
    req.asset_name = asset.name
    req.asset_category = asset.category
    req.asset_serial = asset.serial_number
    req.asset_allocated_date = _ts()
    req.expected_return_date = body.expected_return_date
    req.allocated_by = body.supervisor_name
    req.updated_at = _ts()

    db.add(models.ApprovalHistory(requisition_id=req_id, action="Asset Allocated", action_by=body.supervisor_name, action_role="Supervisor", comment=f"{asset.name} ({body.asset_id})", created_at=_ts()))
    db.commit(); db.refresh(req)
    return req

@app.patch("/requisitions/{req_id}/initiate-return")
def initiate_return(req_id: str, body: RequisitionReturn, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req: raise HTTPException(404, "Requisition not found")
    if req.asset_id:
        asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == req.asset_id).first()
        if asset:
            asset.status = "Return Pending"
            asset.last_updated = _ts()[:10]
    req.status = "return_initiated"
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Return Initiated", action_by=body.initiated_by, action_role="Employee", created_at=_ts()))
    db.commit(); db.refresh(req)
    return req

@app.patch("/requisitions/{req_id}/confirm-return")
def confirm_return(req_id: str, body: RequisitionConfirmReturn, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req: raise HTTPException(404, "Requisition not found")
    if req.asset_id:
        asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == req.asset_id).first()
        if asset:
            asset.status = body.condition
            asset.assigned_to = "Unassigned" if body.condition == "Available" else asset.assigned_to
            asset.assigned_req_id = ""
            asset.last_updated = _ts()[:10]
    req.status = "returned"
    req.actual_return_date = _ts()
    req.return_confirmed_by = body.supervisor_name
    req.return_asset_condition = body.condition
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Return Confirmed", action_by=body.supervisor_name, action_role="Supervisor", comment=f"Condition: {body.condition}", created_at=_ts()))
    db.commit(); db.refresh(req)
    return req

@app.patch("/requisitions/{req_id}/mark-lost")
def mark_lost(req_id: str, body: RequisitionMarkLost, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req: raise HTTPException(404, "Requisition not found")
    if req.asset_id:
        asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == req.asset_id).first()
        if asset:
            asset.status = "Lost"
            asset.last_updated = _ts()[:10]
    req.status = "asset_lost"
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Asset Lost", action_by=body.supervisor_name, action_role="Supervisor", comment=body.notes, created_at=_ts()))
    db.commit(); db.refresh(req)
    return req

@app.get("/requisitions/export/excel")
def export_requisitions_excel(db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    reqs = db.query(models.Requisition).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Requisitions"

    headers = [
        "Requisition ID", "Employee Name", "Employee Department", "Manager Name",
        "Dept Supervisor", "Item Requested", "Quantity", "Reason", "Request Date",
        "Manager Approval Status", "Manager Approval Date", "Manager Rejection Reason",
        "Asset Allocation Status", "Asset Name", "Asset Category", "Asset Serial / ID",
        "Asset Assigned Date", "Expected Return Date", "Actual Return Date",
        "Return Confirmed By", "Return Asset Condition", "Final Status", "Notes",
    ]

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_label = {
        "pending_manager": "Pending Manager Approval",
        "rejected": "Rejected by Manager",
        "manager_approved": "Manager Approved",
        "asset_allocated": "Asset Allocated",
        "return_initiated": "Return Initiated",
        "returned": "Returned & Closed",
        "asset_lost": "Asset Lost",
    }

    for row_idx, r in enumerate(reqs, 2):
        ws.append([
            r.id, r.employee_name, r.employee_dept, r.manager_name or "",
            r.supervisor_name or "", r.item, r.quantity, r.reason,
            r.created_at[:10] if r.created_at else "",
            "Approved" if r.manager_name and r.status != "rejected" else ("Rejected" if r.status == "rejected" else "Pending"),
            r.manager_approval_date[:10] if r.manager_approval_date else "",
            r.rejection_reason or "",
            "Allocated" if r.status in ("asset_allocated", "return_initiated", "returned", "asset_lost") else "",
            r.asset_name or "", r.asset_category or "", r.asset_serial or "",
            r.asset_allocated_date[:10] if r.asset_allocated_date else "",
            r.expected_return_date or "",
            r.actual_return_date[:10] if r.actual_return_date else "",
            r.return_confirmed_by or "", r.return_asset_condition or "",
            status_label.get(r.status, r.status), "",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"requisitions_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Hardware Assets ───────────────────────────────────────────────────────────

class HardwareAssetCreate(BaseModel):
    id: str
    name: str
    category: str
    serial_number: str = ""
    assigned_to: str = "Unassigned"
    dept: str = ""
    location: str = ""
    status: str = "Available"
    purchased: str = ""
    warranty_end: str = ""

@app.get("/hardware-assets")
def list_hardware_assets(db: Session = Depends(get_db)):
    return db.query(models.HardwareAsset).all()

@app.post("/hardware-assets", status_code=201)
def create_hardware_asset(data: HardwareAssetCreate, db: Session = Depends(get_db)):
    asset = models.HardwareAsset(**data.model_dump(), last_updated=_ts()[:10])
    db.add(asset); db.commit(); db.refresh(asset)
    return asset

@app.get("/approval-history/{req_id}")
def get_approval_history(req_id: str, db: Session = Depends(get_db)):
    return db.query(models.ApprovalHistory).filter(models.ApprovalHistory.requisition_id == req_id).all()
