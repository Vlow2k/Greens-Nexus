import io
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
import models
from database import get_db

router = APIRouter(tags=["Requisitions & Hardware Assets"])


def _ts():
    return datetime.utcnow().isoformat()


# ── Schemas ──────────────────────────────────────────────────────────────────

class RequisitionCreate(BaseModel):
    id: str
    employee_name: str
    employee_dept: str
    item: str
    quantity: int = 1
    reason: str = ""
    status: str = "pending_manager"
    supervisor_name: str = ""


class RequisitionApprove(BaseModel):
    manager_name: str


class RequisitionReject(BaseModel):
    manager_name: str
    rejection_reason: str


class RequisitionAllocate(BaseModel):
    asset_id: str
    supervisor_name: str
    expected_return_date: str = ""


class RequisitionReturn(BaseModel):
    initiated_by: str


class RequisitionConfirmReturn(BaseModel):
    supervisor_name: str
    condition: str = "Available"


class RequisitionMarkLost(BaseModel):
    supervisor_name: str
    notes: str = ""


class HardwareAssetCreate(BaseModel):
    id: str
    name: str
    category: str
    serial_number: str = ""
    assigned_to: str = "Unassigned"
    dept: str = ""
    location: str = ""
    status: str = "Available"
    purchased: str = ""
    warranty_end: str = ""


# ── Requisition Routes ────────────────────────────────────────────────────────

@router.get("/requisitions")
def list_requisitions(db: Session = Depends(get_db)):
    return db.query(models.Requisition).all()


@router.post("/requisitions", status_code=201)
def create_requisition(data: RequisitionCreate, db: Session = Depends(get_db)):
    req = models.Requisition(**data.model_dump(), created_at=_ts(), updated_at=_ts())
    db.add(req)
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/approve")
def approve_requisition(req_id: str, body: RequisitionApprove, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    req.status = "manager_approved"
    req.manager_name = body.manager_name
    req.manager_approval_date = _ts()
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Approved", action_by=body.manager_name, action_role="Manager", created_at=_ts()))
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/reject")
def reject_requisition(req_id: str, body: RequisitionReject, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    req.status = "rejected"
    req.manager_name = body.manager_name
    req.rejection_reason = body.rejection_reason
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Rejected", action_by=body.manager_name, action_role="Manager", comment=body.rejection_reason, created_at=_ts()))
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/allocate")
def allocate_asset(req_id: str, body: RequisitionAllocate, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == body.asset_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if not asset:
        raise HTTPException(404, "Asset not found")
    if asset.status != "Available":
        raise HTTPException(400, "Asset not available")

    asset.status = "Checked Out"
    asset.assigned_to = req.employee_name
    asset.assigned_req_id = req_id
    asset.last_updated = _ts()[:10]

    req.status = "asset_allocated"
    req.asset_id = body.asset_id
    req.asset_name = asset.name
    req.asset_category = asset.category
    req.asset_serial = asset.serial_number
    req.asset_allocated_date = _ts()
    req.expected_return_date = body.expected_return_date
    req.allocated_by = body.supervisor_name
    req.updated_at = _ts()

    db.add(models.ApprovalHistory(requisition_id=req_id, action="Asset Allocated", action_by=body.supervisor_name, action_role="Supervisor", comment=f"{asset.name} ({body.asset_id})", created_at=_ts()))
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/initiate-return")
def initiate_return(req_id: str, body: RequisitionReturn, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if req.asset_id:
        asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == req.asset_id).first()
        if asset:
            asset.status = "Return Pending"
            asset.last_updated = _ts()[:10]
    req.status = "return_initiated"
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Return Initiated", action_by=body.initiated_by, action_role="Employee", created_at=_ts()))
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/confirm-return")
def confirm_return(req_id: str, body: RequisitionConfirmReturn, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if req.asset_id:
        asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == req.asset_id).first()
        if asset:
            asset.status = body.condition
            asset.assigned_to = "Unassigned" if body.condition == "Available" else asset.assigned_to
            asset.assigned_req_id = ""
            asset.last_updated = _ts()[:10]
    req.status = "returned"
    req.actual_return_date = _ts()
    req.return_confirmed_by = body.supervisor_name
    req.return_asset_condition = body.condition
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Return Confirmed", action_by=body.supervisor_name, action_role="Supervisor", comment=f"Condition: {body.condition}", created_at=_ts()))
    db.commit()
    db.refresh(req)
    return req


@router.patch("/requisitions/{req_id}/mark-lost")
def mark_lost(req_id: str, body: RequisitionMarkLost, db: Session = Depends(get_db)):
    req = db.query(models.Requisition).filter(models.Requisition.id == req_id).first()
    if not req:
        raise HTTPException(404, "Requisition not found")
    if req.asset_id:
        asset = db.query(models.HardwareAsset).filter(models.HardwareAsset.id == req.asset_id).first()
        if asset:
            asset.status = "Lost"
            asset.last_updated = _ts()[:10]
    req.status = "asset_lost"
    req.updated_at = _ts()
    db.add(models.ApprovalHistory(requisition_id=req_id, action="Asset Lost", action_by=body.supervisor_name, action_role="Supervisor", comment=body.notes, created_at=_ts()))
    db.commit()
    db.refresh(req)
    return req


@router.get("/requisitions/export/excel")
def export_requisitions_excel(db: Session = Depends(get_db)):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed — run: pip install openpyxl")

    reqs = db.query(models.Requisition).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Requisitions"

    headers = [
        "Requisition ID", "Employee Name", "Employee Department", "Manager Name",
        "Dept Supervisor", "Item Requested", "Quantity", "Reason", "Request Date",
        "Manager Approval Status", "Manager Approval Date", "Manager Rejection Reason",
        "Asset Allocation Status", "Asset Name", "Asset Category", "Asset Serial / ID",
        "Asset Assigned Date", "Expected Return Date", "Actual Return Date",
        "Return Confirmed By", "Return Asset Condition", "Final Status", "Notes",
    ]

    header_fill = PatternFill(start_color="1A1A2E", end_color="1A1A2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    status_label = {
        "pending_manager": "Pending Manager Approval",
        "rejected": "Rejected by Manager",
        "manager_approved": "Manager Approved",
        "asset_allocated": "Asset Allocated",
        "return_initiated": "Return Initiated",
        "returned": "Returned & Closed",
        "asset_lost": "Asset Lost",
    }

    for row_idx, r in enumerate(reqs, 2):
        ws.append([
            r.id, r.employee_name, r.employee_dept, r.manager_name or "",
            r.supervisor_name or "", r.item, r.quantity, r.reason,
            r.created_at[:10] if r.created_at else "",
            "Approved" if r.manager_name and r.status != "rejected" else ("Rejected" if r.status == "rejected" else "Pending"),
            r.manager_approval_date[:10] if r.manager_approval_date else "",
            r.rejection_reason or "",
            "Allocated" if r.status in ("asset_allocated", "return_initiated", "returned", "asset_lost") else "",
            r.asset_name or "", r.asset_category or "", r.asset_serial or "",
            r.asset_allocated_date[:10] if r.asset_allocated_date else "",
            r.expected_return_date or "",
            r.actual_return_date[:10] if r.actual_return_date else "",
            r.return_confirmed_by or "", r.return_asset_condition or "",
            status_label.get(r.status, r.status), "",
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"requisitions_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Hardware Assets ───────────────────────────────────────────────────────────

@router.get("/hardware-assets")
def list_hardware_assets(db: Session = Depends(get_db)):
    return db.query(models.HardwareAsset).all()


@router.post("/hardware-assets", status_code=201)
def create_hardware_asset(data: HardwareAssetCreate, db: Session = Depends(get_db)):
    asset = models.HardwareAsset(**data.model_dump(), last_updated=_ts()[:10])
    db.add(asset)
    db.commit()
    db.refresh(asset)
    return asset


@router.get("/approval-history/{req_id}")
def get_approval_history(req_id: str, db: Session = Depends(get_db)):
    return db.query(models.ApprovalHistory).filter(models.ApprovalHistory.requisition_id == req_id).all()
